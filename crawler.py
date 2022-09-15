#-*- encoding: utf8 -*-
import requests
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
import pyperclip
import time
from selenium import webdriver
import re
from openpyxl import Workbook
import sys, os
import math
import psutil


class Naver():
    def __init__(self):
        self.wb = Workbook()
        self.sheet1 = self.wb.active
        self.user_data = []
        time.sleep(1)
        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.driver = webdriver.Chrome(options=options)
    def login(self, id, pw):
        try:
            self.driver.get('https://naver.com/')
            time.sleep(1)
            self.driver.find_element_by_xpath('//*[@id="u_skip"]/a[5]').send_keys(Keys.CONTROL + '\n')
            time.sleep(1)
            self.clipboard_copy('//*[@id="query"]', id)
            time.sleep(1)
            self.driver.find_element_by_xpath('//*[@id="account"]/a').click()
            time.sleep(1)
            self.clipboard_paste('//*[@id="id"]', id)
            time.sleep(1)
            self.driver.switch_to_window(self.driver.window_handles[1])
            time.sleep(1)
            self.clipboard_copy('//*[@id="query"]', pw)
            time.sleep(1)
            self.driver.switch_to_window(self.driver.window_handles[0])
            time.sleep(1)
            self.clipboard_paste('//*[@id="pw"]', pw)
            time.sleep(1)
            self.driver.find_element_by_xpath('//*[@id="log.login"]').click()
            time.sleep(1)
            self.driver.get("https://cafe.naver.com/")
            time.sleep(1)
            #페이지당 100명
            self.driver.find_element_by_xpath('//*[@id="_sortPerPage"]/option[4]').click()
            time.sleep(1)
            return True
        except Exception:
            return False
    def clipboard_paste(self, xpath, word):
        try:
            input_box = self.driver.find_element_by_xpath(xpath)
            input_box.send_keys(Keys.CONTROL, 'v')
            time.sleep(1)
            return True
        except Exception:
            return False
    def clipboard_copy(self, xpath, word):
        try:
            input_box = self.driver.find_element_by_xpath(xpath)
            input_box.send_keys(word)
            input_box.send_keys(Keys.CONTROL, 'a')
            input_box.send_keys(Keys.CONTROL, 'c')
            time.sleep()
            return True
        except Exception:
            return False
    def getToday(self):
        try:
            # 상세검색 클릭
            self.driver.find_element_by_xpath('//*[@id="frmSearch"]/div/div/div[1]/a[2]').click()
            time.sleep(1)
            # 가입/최종방문일 클릭
            self.driver.find_element_by_xpath('//*[@id="srch_date"]').click()
            time.sleep(1)
            # 검색
            self.driver.find_element_by_xpath('//*[@id="frmSearch"]/div/div/div[2]/div[4]/a').click()
            time.sleep(1)
            #페이지 개수
            page_bar = self.driver.find_element_by_xpath('//*[@id="paginate"]')
            pages = page_bar.find_elements_by_css_selector('a')
            page_len = len(pages)
            time.sleep(1)
            for i in range(page_len):
                self.driver.find_element_by_xpath(f'//*[@id="paginate"]/a[{i+1}]').click()
                time.sleep(0.5)
                # 멤버 디테일 보기
                self.driver.find_element_by_xpath('//*[@id="_tableContent"]/table/thead/tr/th[3]/strong/a').click()
                time.sleep(0.4)
                html = self.driver.page_source
                soup = BeautifulSoup(html, 'html.parser')
                data = soup.select("#_tableContent > table > tbody > tr > td > div > div > ol > li > p")
                member_id = soup.select('#_tableContent > table > tbody > tr > td > div > a')
                for j in range(len(data)):
                    if (j+1) % 3 == 0 and data[j].get_text()[0:3] == "010":
                        phone_no = re.sub(r"^(\d{3})[^\d]?(\d{4})[^\d]?(\d{4}).*", r"\1-\2-\3", data[j].get_text())
                        self.user_data.append([member_id[(j+1)//3 - 1].get_text(), phone_no, data[j-2].get_text() + "/" + data[j-1].get_text()])

            return True
        except Exception:
            return False
    def getN(self, n_count):
        # try:
        self.count_data = 0
        html = self.driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        member_count = soup.find('em', class_='_memberCount')
        member_count = int(member_count.get_text().replace(',', ''))
        if member_count < n_count:
            n_count = member_count
        time.sleep(1)
        #페이지 개수
        page_no = math.ceil(n_count / 100)
        for i in range(1, page_no + 1): 
            if i < 11:
                self.driver.find_element_by_xpath(f'//*[@id="paginate"]/a[{i}]').click()
                time.sleep(1)
            else:
                if i%10 == 1:
                    self.driver.find_element_by_css_selector('#paginate > a.next').click()
                    time.sleep(1)
                if i % 10 == 0:
                    self.driver.find_element_by_xpath('//*[@id="paginate"]/a[11]').click()
                else:
                    self.driver.find_element_by_xpath(f'//*[@id="paginate"]/a[{i%10 + 1}]').click()
                time.sleep(1)
            # 1 - 10
            # 1 2 - 11 12
            # 멤버 디테일 보기
            self.driver.find_element_by_xpath('//*[@id="_tableContent"]/table/thead/tr/th[3]/strong/a').click()
            time.sleep(0.4)
            html = self.driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            data = soup.select("#_tableContent > table > tbody > tr > td > div > div > ol > li > p")
            member_id = soup.select('#_tableContent > table > tbody > tr > td > div > a')
            for j in range(len(data)):
                if self.count_data == n_count:
                    return True
                if (j+1) % 3 == 0:
                    if data[j].get_text()[0:3] == "010":
                        phone_no = re.sub(r"^(\d{3})[^\d]?(\d{4})[^\d]?(\d{4}).*", r"\1-\2-\3", data[j].get_text())
                        self.user_data.append([member_id[(j+1)//3 - 1].get_text(), phone_no, data[j-2].get_text() + "/" + data[j-1].get_text()])
                    self.count_data += 1
        return True
        # except Exception as e:
        #     print(self.count_data)
        #     print(e)
        #     return False
    def saveExcel(self, excel_file_name):
        try:
            for i in range(len(self.user_data)):
                self.sheet1.cell(row=i+1, column=1).value = self.user_data[i][0]
                self.sheet1.cell(row=i+1, column=3).value = self.user_data[i][1]
                self.sheet1.cell(row=i+1, column=5).value = self.user_data[i][2]
            self.wb.save(filename=excel_file_name+".xlsx")
            return True
        except Exception:
            return False


if __name__ == "__main__":
    # memory_usage_dict = dict(psutil.virtual_memory()._asdict())
    # memory_usage_percent = memory_usage_dict['percent']
    # print(f"before memory_usage_percent: {memory_usage_percent}%")
    # # current process RAM usage
    # pid = os.getpid()
    # current_process = psutil.Process(pid)
    # current_process_memory_usage_as_KB = current_process.memory_info()[0] / 2.**20
    # print(f"before Current memory KB   : {current_process_memory_usage_as_KB: 9.3f} KB")

    search_option = input("1. 오늘 가입자 정보 가져오기\n2. 최근 N명 정보 가져오기\n입력 : ")
    if search_option != "1" and search_option != "2":
        sys.exit("Error")
    if search_option == "2":
        n_count = int(input("가져올 사람 수를 입력하세요 : "))
    code = input()
    naver = Naver()
    naver.login(code)
    if search_option == "1":
        naver.getToday()
    elif search_option == "2":
        naver.getN(n_count)
    naver.saveExcel("zzzwd")

    # memory_usage_dict = dict(psutil.virtual_memory()._asdict())
    # memory_usage_percent = memory_usage_dict['percent']
    # print(f"after memory_usage_percent: {memory_usage_percent}%")
    # # current process RAM usage
    # pid = os.getpid()
    # current_process = psutil.Process(pid)
    # current_process_memory_usage_as_KB = current_process.memory_info()[0] / 2.**20
    # print(f"after Current memory KB   : {current_process_memory_usage_as_KB: 9.3f} KB")





# 멤버관리창으로 


# headers = {
#     'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'
# }
# s = requests.Session()
# s.headers.update(headers)

# for cookie in driver.get_cookies():
#     c = {cookie['name'] : cookie['value']}
#     s.cookies.update(c)
# res = s.get("https://cafe.naver.com/ManageWholeMember.nhn?clubid=29470508")
# soup = BeautifulSoup(res.text, 'lxml')
# print(soup.prettify())